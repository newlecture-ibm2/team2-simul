#!/usr/bin/env python3
"""Simul WBS Excel Generator — Gantt chart with elegant design."""

import csv
import math
import os
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── Configuration ──────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
CSV_PATH = os.path.join(SCRIPT_DIR, "simul_wbs.csv")
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "Simul_WBS.xlsx")

PROJECT_START = date(2026, 4, 27)
PROJECT_END = date(2026, 5, 29)
HOURS_PER_DAY = 8

# ── Color Palette ──────────────────────────────────────────────
HEADER_BG = "1B1F3B"
HEADER_FONT_COLOR = "FFFFFF"
BORDER_LIGHT = "E0E0E0"
ROW_EVEN = "F8F9FA"
ROW_ODD = "FFFFFF"
WEEKEND_BG = "F0F0F5"

EPIC_COLORS = {
    "공통 인프라":   {"bg": "2D3250", "bar": "6C63FF"},
    "인증/사용자":   {"bg": "2D3250", "bar": "FF6B6B"},
    "AI 가상시착":   {"bg": "2D3250", "bar": "4ECDC4"},
    "개인 옷장":     {"bg": "2D3250", "bar": "FFD93D"},
    "커뮤니티 피드":  {"bg": "2D3250", "bar": "6BCB77"},
    "태그/검색":     {"bg": "2D3250", "bar": "FF8E53"},
    "알림":         {"bg": "2D3250", "bar": "A78BFA"},
    "관리자":        {"bg": "2D3250", "bar": "F87171"},
    "QA/통합":      {"bg": "2D3250", "bar": "38BDF8"},
}

# ── Styles ─────────────────────────────────────────────────────
thin_border = Border(
    left=Side(style='thin', color=BORDER_LIGHT),
    right=Side(style='thin', color=BORDER_LIGHT),
    top=Side(style='thin', color=BORDER_LIGHT),
    bottom=Side(style='thin', color=BORDER_LIGHT),
)
header_border = Border(
    left=Side(style='thin', color="3D3D5C"),
    right=Side(style='thin', color="3D3D5C"),
    top=Side(style='thin', color="3D3D5C"),
    bottom=Side(style='medium', color="3D3D5C"),
)

# ── Parse CSV ──────────────────────────────────────────────────
def parse_csv(path):
    epics, tasks = [], []
    with open(path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            (epics if row['Issue Type'] == 'Epic' else tasks).append(row)
    return epics, tasks

# ── Schedule tasks ─────────────────────────────────────────────
def schedule_tasks(tasks):
    assignee_next = {}
    scheduled = []
    for t in tasks:
        assignee = t.get('Assignee', '').strip('"')
        hours_str = t.get('Original Estimate', '0h').replace('h', '').strip()
        try:
            hours = int(hours_str)
        except ValueError:
            hours = 8
        days_needed = max(1, math.ceil(hours / HOURS_PER_DAY))
        start = assignee_next.get(assignee, PROJECT_START)
        while start.weekday() >= 5:
            start += timedelta(days=1)
        end = start
        work_days = 1
        while work_days < days_needed:
            end += timedelta(days=1)
            if end.weekday() < 5:
                work_days += 1
        if end > PROJECT_END:
            end = PROJECT_END
        nxt = end + timedelta(days=1)
        while nxt.weekday() >= 5:
            nxt += timedelta(days=1)
        assignee_next[assignee] = nxt
        scheduled.append({**t, 'start_date': start, 'end_date': end, 'days': days_needed, 'hours': hours})
    return scheduled

# ── Build Excel ────────────────────────────────────────────────
def build_excel(epics, scheduled_tasks):
    wb = Workbook()
    ws = wb.active
    ws.title = "Simul WBS"

    fixed_cols = [
        ("No.", 5), ("WBS ID", 14), ("Epic", 12), ("구분", 10),
        ("태스크명", 42), ("우선순위", 8), ("MoSCoW", 8),
        ("담당자", 9), ("공수(h)", 7), ("시작일", 11),
        ("종료일", 11), ("기간(일)", 7), ("진행률", 7),
    ]
    num_fixed = len(fixed_cols)
    total_days = (PROJECT_END - PROJECT_START).days + 1
    all_dates = [PROJECT_START + timedelta(days=i) for i in range(total_days)]

    for i, (_, w) in enumerate(fixed_cols, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for i in range(total_days):
        ws.column_dimensions[get_column_letter(num_fixed + 1 + i)].width = 3.2

    # Row 1: Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_fixed + total_days)
    tc = ws.cell(row=1, column=1, value="Simul — AI 패션 소셜 플랫폼 MVP WBS")
    tc.font = Font(name='맑은 고딕', size=16, bold=True, color="1B1F3B")
    tc.alignment = Alignment(horizontal='center', vertical='center')
    tc.fill = PatternFill(start_color="F4F4FB", end_color="F4F4FB", fill_type="solid")
    ws.row_dimensions[1].height = 36

    # Row 2: Info + month headers
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_fixed)
    ic = ws.cell(row=2, column=1,
                 value=f"프로젝트 기간: {PROJECT_START.strftime('%Y.%m.%d')} ~ {PROJECT_END.strftime('%Y.%m.%d')}  |  일 {HOURS_PER_DAY}시간 기준")
    ic.font = Font(name='맑은 고딕', size=9, color="666666")
    ic.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 22

    current_month = None
    month_start_col = None
    for i, d in enumerate(all_dates):
        col = num_fixed + 1 + i
        mk = (d.year, d.month)
        if mk != current_month:
            if current_month and month_start_col:
                ws.merge_cells(start_row=2, start_column=month_start_col, end_row=2, end_column=col - 1)
                mc = ws.cell(row=2, column=month_start_col, value=f"{current_month[0]}.{current_month[1]:02d}")
                mc.font = Font(name='맑은 고딕', size=9, bold=True, color="FFFFFF")
                mc.fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
                mc.alignment = Alignment(horizontal='center', vertical='center')
            current_month = mk
            month_start_col = col
    if month_start_col:
        ws.merge_cells(start_row=2, start_column=month_start_col, end_row=2, end_column=num_fixed + total_days)
        mc = ws.cell(row=2, column=month_start_col, value=f"{current_month[0]}.{current_month[1]:02d}")
        mc.font = Font(name='맑은 고딕', size=9, bold=True, color="FFFFFF")
        mc.fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
        mc.alignment = Alignment(horizontal='center', vertical='center')

    # Row 3: Column headers + day numbers
    ws.row_dimensions[3].height = 28
    for i, (name, _) in enumerate(fixed_cols, 1):
        c = ws.cell(row=3, column=i, value=name)
        c.font = Font(name='맑은 고딕', size=9, bold=True, color=HEADER_FONT_COLOR)
        c.fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = header_border
    for i, d in enumerate(all_dates):
        col = num_fixed + 1 + i
        c = ws.cell(row=3, column=col, value=d.day)
        is_wk = d.weekday() >= 5
        c.font = Font(name='맑은 고딕', size=7, bold=True, color="FF6B6B" if is_wk else HEADER_FONT_COLOR)
        bg = "3D3250" if is_wk else HEADER_BG
        c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = header_border

    # Row 4: Day-of-week
    day_kr = ['월','화','수','목','금','토','일']
    ws.row_dimensions[4].height = 16
    for i in range(1, num_fixed + 1):
        c = ws.cell(row=4, column=i)
        c.fill = PatternFill(start_color="2D3250", end_color="2D3250", fill_type="solid")
        c.border = header_border
    for i, d in enumerate(all_dates):
        col = num_fixed + 1 + i
        c = ws.cell(row=4, column=col, value=day_kr[d.weekday()])
        is_wk = d.weekday() >= 5
        c.font = Font(name='맑은 고딕', size=7, color="FF6B6B" if is_wk else "AAAACC")
        bg = "3D3250" if is_wk else "2D3250"
        c.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = header_border

    # Data rows
    row_idx = 5
    task_no = 0
    epic_order = [e['Epic Name'] for e in epics]
    tasks_by_epic = {}
    for t in scheduled_tasks:
        ep = t.get('Epic Link', '').strip()
        tasks_by_epic.setdefault(ep, []).append(t)

    for epic_name in epic_order:
        epic_tasks = tasks_by_epic.get(epic_name, [])
        if not epic_tasks:
            continue
        colors = EPIC_COLORS.get(epic_name, {"bg": "2D3250", "bar": "888888"})

        # Epic row
        ws.row_dimensions[row_idx].height = 26
        ef = PatternFill(start_color=colors["bg"], end_color=colors["bg"], fill_type="solid")
        efont = Font(name='맑은 고딕', size=10, bold=True, color="FFFFFF")
        for col in range(1, num_fixed + total_days + 1):
            ws.cell(row=row_idx, column=col).fill = ef
            ws.cell(row=row_idx, column=col).border = thin_border
        ws.cell(row=row_idx, column=3, value=epic_name).font = efont
        ws.cell(row=row_idx, column=3).fill = ef
        ws.cell(row=row_idx, column=3).alignment = Alignment(horizontal='left', vertical='center')
        total_h = sum(t['hours'] for t in epic_tasks)
        hc = ws.cell(row=row_idx, column=9, value=f"{total_h}h")
        hc.font = Font(name='맑은 고딕', size=9, bold=True, color="FFFFFF")
        hc.fill = ef
        hc.alignment = Alignment(horizontal='center', vertical='center')
        # Epic bar
        e_start = min(t['start_date'] for t in epic_tasks)
        e_end = max(t['end_date'] for t in epic_tasks)
        for i, d in enumerate(all_dates):
            col = num_fixed + 1 + i
            c = ws.cell(row=row_idx, column=col)
            if e_start <= d <= e_end:
                c.fill = PatternFill(start_color=colors["bar"], end_color=colors["bar"], fill_type="solid")
        row_idx += 1

        # Task rows
        for ti, t in enumerate(epic_tasks):
            task_no += 1
            ws.row_dimensions[row_idx].height = 22
            row_bg = ROW_EVEN if ti % 2 == 0 else ROW_ODD
            rf = PatternFill(start_color=row_bg, end_color=row_bg, fill_type="solid")
            summary = t.get('Summary', '')
            wbs_id, task_name = "", summary
            if summary.startswith('['):
                be = summary.find(']')
                if be > 0:
                    wbs_id = summary[1:be]
                    task_name = summary[be+1:].strip()
            assignee = t.get('Assignee', '').strip('"')
            priority = t.get('Priority', '')
            moscow = t.get('Labels', '')
            component = t.get('Component/s', '')

            vals = [task_no, wbs_id, epic_name, component, task_name, priority, moscow,
                    assignee, t['hours'], t['start_date'].strftime('%m/%d'),
                    t['end_date'].strftime('%m/%d'), t['days'], "0%"]
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=row_idx, column=ci, value=v)
                c.fill = rf
                c.border = thin_border
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                c.font = Font(name='맑은 고딕', size=8.5, color="333333")
                if ci == 2:
                    c.font = Font(name='Consolas', size=8, bold=True, color=colors["bar"])
                elif ci == 5:
                    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
                    c.font = Font(name='맑은 고딕', size=8.5, color="1B1F3B")
                elif ci == 6:
                    clr = {"Highest":"FF4444","High":"FF8800","Medium":"FFAA00"}.get(priority,"333333")
                    c.font = Font(name='맑은 고딕', size=8, bold=priority in ('Highest','High'), color=clr)
                elif ci == 7:
                    clr = {"Must":"D32F2F","Should":"F57C00","Could":"1976D2"}.get(moscow,"333333")
                    c.font = Font(name='맑은 고딕', size=8, bold=moscow=='Must', color=clr)

            # Gantt bar
            for i, d in enumerate(all_dates):
                col = num_fixed + 1 + i
                c = ws.cell(row=row_idx, column=col)
                is_wk = d.weekday() >= 5
                in_range = t['start_date'] <= d <= t['end_date']
                if in_range and not is_wk:
                    c.fill = PatternFill(start_color=colors["bar"], end_color=colors["bar"], fill_type="solid")
                elif is_wk:
                    c.fill = PatternFill(start_color=WEEKEND_BG, end_color=WEEKEND_BG, fill_type="solid")
                    c.border = thin_border
                else:
                    c.fill = rf
                    c.border = thin_border
            row_idx += 1

    # Summary row
    ws.row_dimensions[row_idx].height = 28
    sf = PatternFill(start_color="1B1F3B", end_color="1B1F3B", fill_type="solid")
    sfont = Font(name='맑은 고딕', size=10, bold=True, color="FFFFFF")
    for col in range(1, num_fixed + total_days + 1):
        ws.cell(row=row_idx, column=col).fill = sf
        ws.cell(row=row_idx, column=col).border = thin_border
    ws.cell(row=row_idx, column=1, value=f"{task_no}").font = sfont
    ws.cell(row=row_idx, column=1).fill = sf
    ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row_idx, column=5, value="TOTAL").font = sfont
    ws.cell(row=row_idx, column=5).fill = sf
    ws.cell(row=row_idx, column=5).alignment = Alignment(horizontal='center', vertical='center')
    th = sum(t['hours'] for t in scheduled_tasks)
    ws.cell(row=row_idx, column=9, value=f"{th}h").font = sfont
    ws.cell(row=row_idx, column=9).fill = sf
    ws.cell(row=row_idx, column=9).alignment = Alignment(horizontal='center', vertical='center')

    # Freeze panes
    ws.freeze_panes = f'{get_column_letter(num_fixed + 1)}5'
    ws.page_setup.orientation = 'landscape'

    # Legend sheet
    ws2 = wb.create_sheet("범례")
    ws2.column_dimensions['A'].width = 15
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 40
    ws2.cell(row=1, column=1, value="범례").font = Font(name='맑은 고딕', size=14, bold=True, color="1B1F3B")
    ws2.merge_cells('A1:C1')
    lh = Font(name='맑은 고딕', size=10, bold=True, color="FFFFFF")
    lf = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
    for ci, h in enumerate(["Epic", "컬러", "담당자"], 1):
        c = ws2.cell(row=3, column=ci, value=h)
        c.font = lh; c.fill = lf; c.alignment = Alignment(horizontal='center', vertical='center')
    for ri, ep in enumerate(epics, 4):
        nm = ep['Epic Name']
        clr = EPIC_COLORS.get(nm, {"bar": "888888"})
        ws2.cell(row=ri, column=1, value=nm).font = Font(name='맑은 고딕', size=9, bold=True)
        ws2.cell(row=ri, column=2, value="████████").font = Font(name='맑은 고딕', size=10, color=clr["bar"])
        ws2.cell(row=ri, column=3, value=ep.get('Assignee','').strip('"')).font = Font(name='맑은 고딕', size=9)
    r = len(epics) + 6
    ws2.cell(row=r, column=1, value="MoSCoW 분류").font = Font(name='맑은 고딕', size=12, bold=True, color="1B1F3B")
    for ri, (lb, ds) in enumerate([("Must","MVP 필수 구현"),("Should","중요하지만 여력 시"),("Could","MVP 이후 추가 가능")], r+1):
        ws2.cell(row=ri, column=1, value=lb).font = Font(name='맑은 고딕', size=9, bold=True)
        ws2.cell(row=ri, column=2, value=ds).font = Font(name='맑은 고딕', size=9)
    return wb

def main():
    print("Simul WBS Excel 생성 중...")
    epics, tasks = parse_csv(CSV_PATH)
    print(f"  CSV 파싱 완료: Epic {len(epics)}개, Task {len(tasks)}개")
    scheduled = schedule_tasks(tasks)
    print(f"  스케줄링 완료: {len(scheduled)}개 태스크")
    total_h = sum(t['hours'] for t in scheduled)
    print(f"  총 공수: {total_h}h")
    wb = build_excel(epics, scheduled)
    wb.save(OUTPUT_PATH)
    print(f"  저장 완료: {OUTPUT_PATH}")
    print("Done!")

if __name__ == "__main__":
    main()
